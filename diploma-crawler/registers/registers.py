import urllib.request, requests, json, bs4, webbrowser
from openpyxl import load_workbook
from pymongo import MongoClient
from time import sleep


def update_registers():
    client = MongoClient("mongodb://admin:myadminpassword@217.73.60.165:2327/")
    db = client.registers_CB

    MFO = db.MFO
    MFO.delete_many({})
    outfilename = "list_MFO.xlsx"
    url_of_file = "http://www.cbr.ru/vfs/finmarkets/files/supervision/"
    urllib.request.urlretrieve(url_of_file + outfilename, outfilename)
    wb_val = load_workbook(filename=outfilename, data_only=True)
    sheet_obj = wb_val.worksheets[0]
    for i in range(6, sheet_obj.max_row + 1):
        h = {'number ': sheet_obj.cell(row=i, column=1).value,
             'reg_num': sheet_obj.cell(row=i, column=2).value + sheet_obj.cell(row=i, column=3).value + sheet_obj.cell(
                 row=i, column=4).value + sheet_obj.cell(row=i, column=5).value + sheet_obj.cell(row=i, column=6).value,
             'date': sheet_obj.cell(row=i, column=7).value,
             'form_num': sheet_obj.cell(row=i, column=8).value,
             'type': sheet_obj.cell(row=i, column=9).value,
             'OGRN': sheet_obj.cell(row=i, column=10).value,
             'INN': sheet_obj.cell(row=i, column=11).value,
             'full_name': sheet_obj.cell(row=i, column=12).value,
             'name': sheet_obj.cell(row=i, column=13).value,
             'address': sheet_obj.cell(row=i, column=14).value,
             'website': sheet_obj.cell(row=i, column=15).value,
             'email': sheet_obj.cell(row=i, column=16).value}
        MFO.insert_one(h)

    SRO = db.SRO
    SRO.delete_many({})
    SRO_members = db.SRO_members
    SRO_members.delete_many({})
    outfilename = "list_sro.xlsx"
    url_of_file = "http://www.cbr.ru/vfs/finmarkets/files/supervision/"
    urllib.request.urlretrieve(url_of_file + outfilename, outfilename)
    wb_val = load_workbook(filename=outfilename, data_only=True)
    sheet_obj = wb_val.worksheets[0]
    for i in range(7, sheet_obj.max_row + 1):
        if sheet_obj.cell(row=i, column=1).value:
            h = {'number ': sheet_obj.cell(row=i, column=1).value,
                 'date': sheet_obj.cell(row=i, column=2).value,
                 'full_name': sheet_obj.cell(row=i, column=3).value,
                 'legal_form': sheet_obj.cell(row=i, column=4).value,
                 'INN': sheet_obj.cell(row=i, column=5).value,
                 'OGRN': sheet_obj.cell(row=i, column=6).value,
                 'address': sheet_obj.cell(row=i, column=7).value,
                 'website': sheet_obj.cell(row=i, column=8).value,
                 'phone': sheet_obj.cell(row=i, column=9).value,
                 'email': sheet_obj.cell(row=i, column=10).value}
            SRO.insert_one(h)
            sheet_obj1 = wb_val[str(sheet_obj.cell(row=i, column=1).value)]
            for j in range(7, sheet_obj1.max_row + 1):
                h = {'SRO_num': sheet_obj.cell(row=i, column=1).value,
                     'SRO_name': sheet_obj.cell(row=i, column=3).value,
                     'number ': sheet_obj1.cell(row=j, column=1).value,
                     'full_name': sheet_obj1.cell(row=j, column=2).value,
                     'INN': sheet_obj1.cell(row=j, column=3).value,
                     'OGRN': sheet_obj1.cell(row=j, column=4).value,
                     'type': sheet_obj1.cell(row=j, column=5).value}
                SRO_members.insert_one(h)

    AIF = db.AIF
    AIF.delete_many({})
    outfilename = "list_if.xlsx"
    url_of_file = "https://cbr.ru/vfs/finmarkets/files/supervision/"
    urllib.request.urlretrieve(url_of_file + outfilename, outfilename)
    wb_val = load_workbook(filename=outfilename, data_only=True)
    sheet_obj = wb_val.worksheets[0]
    for i in range(2, sheet_obj.max_row + 1):
        h = {'number ': sheet_obj.cell(row=i, column=1).value,
             'full_name': sheet_obj.cell(row=i, column=2).value,
             'name': sheet_obj.cell(row=i, column=3).value,
             'INN': sheet_obj.cell(row=i, column=4).value,
             'license': sheet_obj.cell(row=i, column=5).value,
             'date': sheet_obj.cell(row=i, column=6).value,
             'validity': sheet_obj.cell(row=i, column=7).value,
             'address': sheet_obj.cell(row=i, column=8).value,
             'phone': sheet_obj.cell(row=i, column=9).value,
             'website': sheet_obj.cell(row=i, column=10).value, }
        AIF.insert_one(h)

    OPS = db.OPS
    OPS.delete_many({})
    outfilename = "list_ops.xlsx"
    url_of_file = "https://cbr.ru/vfs/PSystem/rops/"
    urllib.request.urlretrieve(url_of_file + outfilename, outfilename)
    wb_val = load_workbook(filename=outfilename, data_only=True)
    sheet_obj = wb_val.worksheets[0]
    column = ['', 'number', 'reg_num', 'dt_reg', 'OGRN', '', 'reg_num_ops', 'full_name', 'address', 'name_ops',
              'sc_name', 'sc_addres', 'sc_OGRN', 'sc_reg_num', 'pcc_name', 'pcc_address', 'pcc_OGRM', '', 'pcc_reg_num',
              'pcc_code', 'oc_name', 'oc_address', 'oc_OGRM', '', 'oc_reg_num', 'oc_code', 'importance', 'start_dt',
              'end_dt', 'end_info']
    for i in range(7, sheet_obj.max_row + 1):
        h = {}
        for j in range(1, 29):
            k = 0
            y = 0
            if (j == 4 or j == 16 or j == 22):
                while (sheet_obj.cell(row=i - k, column=j).value is None and sheet_obj.cell(row=i - k,
                                                                                            column=j + 1).value is None):
                    k = k + 1
                if (sheet_obj.cell(row=i - k, column=j).value is None):
                    y = 1
                h.update({column[j]: sheet_obj.cell(row=i - k, column=j + y).value})
                j = j + 1
            elif (j != 5 and j != 17 and j != 23):
                while (sheet_obj.cell(row=i - k, column=j).value is None):
                    k = k + 1
                h.update({column[j]: sheet_obj.cell(row=i - k, column=j).value})
        OPS.insert_one(h)

    FOREX = db.FOREX
    FOREX.delete_many({})
    outfilename = "list_forex_dealers.xlsx"
    url_of_file = "http://www.cbr.ru/vfs/finmarkets/files/supervision/"
    urllib.request.urlretrieve(url_of_file + outfilename, outfilename)
    wb_val = load_workbook(filename=outfilename, data_only=True)
    sheet_obj = wb_val.worksheets[0]
    for i in range(5, sheet_obj.max_row + 1):
        h = {'number ': sheet_obj.cell(row=i, column=1).value,
             'name': sheet_obj.cell(row=i, column=2).value,
             'INN': sheet_obj.cell(row=i, column=3).value,
             'OGRN': sheet_obj.cell(row=i, column=4).value,
             'address': sheet_obj.cell(row=i, column=5).value,
             'phone': sheet_obj.cell(row=i, column=6).value,
             'license': sheet_obj.cell(row=i, column=7).value,
             'date': sheet_obj.cell(row=i, column=8).value,
             'validity': sheet_obj.cell(row=i, column=9).value,
             'status': sheet_obj.cell(row=i, column=10).value,
             'region': sheet_obj.cell(row=i, column=11).value,
             'supervision': sheet_obj.cell(row=i, column=12).value}
        FOREX.insert_one(h)

    return update_registers

if __name__ == "__main__":
    update_registers()
